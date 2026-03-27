# pip install beautifulsoup4 openpyxl fake_useragent selenium
# Race Meetings.xlsx

import re
import time
import os
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager

ChromeDriverPath = "C:/chromedriver/chromedriver.exe"

BASE_URL = 'https://www.tab.com.au'
FILE_NAME = 'Race Meetings.xlsm'
target_column = 23
ALLOWED_MEETINGS = ['(VIC)', '(NSW)', '(QLD)', '(SA)', '(WA)', '(NT)', '(TAS)', '(ACT)', '(NZ)', '(NZL)']
FS = {}
SR = {}

def _normalize_meeting_name(name: str) -> str:
    s = str(name)
    # Drop state/country suffixes like "(QLD)" if present.
    s = re.sub(r"\s*\(.*?\)\s*", " ", s)
    s = s.strip().lower().replace("-", " ")
    s = re.sub(r"\s+", " ", s)
    return s

def _cell_value_with_merges(ws, cell_addr: str):
    """
    If cell_addr is within a merged range, return the top-left cell's value.
    Otherwise return the cell's own value.
    """
    try:
        cell = ws[cell_addr]
    except Exception:
        return None

    try:
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(row=merged.min_row, column=merged.min_col).value
    except Exception:
        pass

    return cell.value

def get_target_meetings_from_excel(excel_file: str) -> set[str] | None:
    """
    Reads meeting/track names from cell G1 of *every* sheet in the workbook.
    Returns a set of normalized meeting names (lowercase, '-' -> ' ', removes '(QLD)' suffixes).
    Only sheets with a non-empty G1 are included. If none found, returns None and the scraper
    will process all meetings (current behavior).
    """
    try:
        wb = load_workbook(filename=excel_file, keep_vba=True, data_only=True, read_only=True)
    except Exception as e:
        print(f"WARNING: Could not open workbook '{excel_file}' to read G1: {e}")
        return None

    try:
        targets: set[str] = set()
        for ws in wb.worksheets:
            raw = _cell_value_with_merges(ws, "G1")
            if raw is None:
                continue

            val = str(raw).strip()
            if not val:
                continue

            targets.add(_normalize_meeting_name(val))
        return targets or None
    finally:
        try:
            wb.close()
        except Exception:
            pass

def _dump_debug(driver, prefix: str):
    try:
        driver.save_screenshot(f"{prefix}.png")
    except Exception:
        pass
    try:
        with open(f"{prefix}.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source or "")
    except Exception:
        pass

def _create_chrome_service():
    """
    Prefer an explicitly configured chromedriver path when present.
    Otherwise fall back to webdriver_manager (auto-download).
    If that fails (e.g., offline/proxy), let Selenium Manager try.
    """
    configured = os.environ.get("CHROMEDRIVER_PATH") or ChromeDriverPath
    if configured:
        p = Path(configured)
        if p.is_file():
            return Service(str(p))

    try:
        return Service(ChromeDriverManager().install())
    except Exception:
        return None

def setup_driver():
    options = Options()
    headless_env = os.environ.get("HEADLESS", "1").strip().lower()
    headless = headless_env not in {"0", "false", "no", "off"}
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-images")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--lang=en-AU,en")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("--window-size=1920,1080")

    # Reduce obvious automation fingerprints
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    # Optional: run using a real Chrome profile (often helps bypass 403/bot blocking)
    user_data_dir = os.environ.get("CHROME_USER_DATA_DIR")
    if user_data_dir:
        options.add_argument(f"--user-data-dir={user_data_dir}")
    profile_dir = os.environ.get("CHROME_PROFILE_DIR")
    if profile_dir:
        options.add_argument(f"--profile-directory={profile_dir}")

    # Optional: ignore cert errors if your network MITMs TLS (use only if needed)
    ignore_cert_env = os.environ.get("IGNORE_CERT_ERRORS", "0").strip().lower()
    if ignore_cert_env in {"1", "true", "yes", "on"}:
        options.add_argument("--ignore-certificate-errors")
        options.add_argument("--allow-insecure-localhost")

    service = _create_chrome_service()
    if service is None:
        driver = webdriver.Chrome(options=options)
    else:
        driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(800)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    return driver

def find_all_races(html):
    soup = BeautifulSoup(html, 'html.parser')
    race_links = soup.select("a[data-testid='race']")
    rounds_links = [a.get("href") for a in race_links if a.get("href")]

    meeting_names = set()
    for href in rounds_links:
        parts = str(href).split("/")
        if len(parts) > 3:
            meeting_names.add(_normalize_meeting_name(parts[3]))

    meetings_names = sorted(meeting_names)
    print("Meetings found (from race links):", meetings_names)
    return meetings_names, rounds_links

def extract_sky_rating(driver, url, allowed_meetings: set[str] | None):
    global SR
    meeting_name = url.split('/')[3]

    if (allowed_meetings is None) or (_normalize_meeting_name(meeting_name) in allowed_meetings):
        SR.setdefault(meeting_name, {})
        try:
            driver.get(BASE_URL + url)
        except:
            driver.execute_script("window.stop()")

        time.sleep(5)
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")

        # Each horse row
        rows = soup.select("div.row")  # Debug print

        for row in rows:
            try:
                horse_el = row.select_one("div.runner-name")
                if not horse_el:
                    continue

                horse_name = horse_el.get_text(strip=True).split("(")[0].strip()
                # Sky Rating is inside a <div> with a numeric value
                rating_el = row.select_one("div.runner-rating-cell span")
                if rating_el:
                    sky_rating = rating_el.get_text(strip=True)

                    if sky_rating.isdigit():
                        SR[meeting_name][horse_name] = sky_rating
                        print("Sky Ratinggggggggggggggggg:", meeting_name, horse_name, sky_rating)

            except Exception as e:
                print(f"Error extracting horse data: {e}")  # Log any errors
                continue

def extract_FS(driver, url, allowed_meetings: set[str] | None):
    global FS
    meeting_name = url.split('/')[3]
    if (allowed_meetings is None) or (_normalize_meeting_name(meeting_name) in allowed_meetings):
        try:
            driver.get(BASE_URL + url)
        except:
            driver.execute_script("window.stop()")
        try:
            if FS[meeting_name]:
                pass
        except:
            FS[meeting_name] = {}
        try:
            button = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Show All Form']]")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
            button.click()
        except:
            pass
        while True:
            time.sleep(1)
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            FS_element = soup.find_all('p', {'class': 'comment-paragraph'})
            horse_name_divs = soup.find_all('div', {'class': 'row active'})
            if FS_element.__len__() > 0:
                print(FS_element.__len__(), horse_name_divs.__len__())
                for i in range(FS_element.__len__()):
                    horse_name = BeautifulSoup(str(horse_name_divs[i]), 'html.parser').find('div', {'class': 'runner-name'}).text.split('(')[0].strip()
                    FS[meeting_name][horse_name] = re.search(r"\(([-+]?\d*\.?\d+)\)", FS_element[i].text).group(1)
                print(FS)
                break


def get_meetings(driver, url, target_meetings: set[str] | None = None):
    def _load(u: str):
        try:
            driver.get(u)
            return True
        except Exception:
            try:
                driver.execute_script("window.stop()")
            except Exception:
                pass
            return False

    # TAB sometimes redirects or fails; retry a few times.
    _load(url)
    _load(url + "R")
    _load(url + "R")

    # WAIT for meeting cards
    try:
        WebDriverWait(driver, 60).until(
            EC.any_of(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-testid='meeting']")),
                EC.presence_of_element_located((By.CSS_SELECTOR, "a[data-testid='race']")),
            )
        )
    except Exception as e:
        print(" ERROR: TAB meetings did not load.")
        print(f"  URL: {getattr(driver, 'current_url', None)}")
        print(f"  Title: {getattr(driver, 'title', None)}")
        print(f"  Wait error: {e}")
        _dump_debug(driver, "debug_tab_meetings")
        print("  Saved debug files: debug_tab_meetings.png, debug_tab_meetings.html")
        return

    html = driver.page_source
    meetings_names, rounds_links = find_all_races(html=html)

    if target_meetings:
        available = set(meetings_names)
        found = sorted(target_meetings.intersection(available))
        missing = sorted(target_meetings.difference(available))

        print(f"[Target] Meetings from Excel (all sheets G1): {sorted(target_meetings)}")
        print(f"[Found] Target meetings found on TAB page: {found}")
        print(f"[Missing] Target meetings missing from TAB page: {missing}")

        if not found:
            print("WARNING: No target meetings found on TAB page, skipping scrape.")
            return

        target_meetings = set(found)

        def _href_meeting_norm(href: str) -> str:
            parts = str(href).split("/")
            # Expected: /racing/meeting/<meeting-slug>/race/<n> ...
            if len(parts) > 3:
                return _normalize_meeting_name(parts[3])
            return ""

        before = len(rounds_links)
        rounds_links = [h for h in rounds_links if _href_meeting_norm(h) in target_meetings]
        print(f"[Filter] Race links: {before} -> {len(rounds_links)}")

    allowed_meetings = target_meetings


    for i in range(rounds_links.__len__()):
        extract_FS(driver, rounds_links[i], allowed_meetings)
        extract_sky_rating(driver, rounds_links[i], allowed_meetings)


def merge_excel(excel_file, FS):
    print("\n==============================")
    print("DEBUG: Starting merge_excel")
    print("==============================")

    workbook = load_workbook(filename=excel_file, keep_vba=True)

    def normalize(name):
        return name.strip().lower().replace("-", " ")

    # Normalize all sheet names
    normalized_sheet_map = {normalize(name): name for name in workbook.sheetnames}

    print("\nSheets in workbook:")
    for k, v in normalized_sheet_map.items():
        print(f"  '{k}'  ->  '{v}'")

    print("\nFS meetings loaded:", list(FS.keys()))
    print("SR meetings loaded:", list(SR.keys()))

    # --- PROCESS FS (TAB FS) ---
    print("\n==============================")
    print("PROCESSING TAB FS (Col W)")
    print("==============================")

    for raw_sheet_name, horses in FS.items():
        norm_name = normalize(raw_sheet_name)
        actual_sheet_name = normalized_sheet_map.get(norm_name)

        print(f"\n➡ Meeting FS: '{raw_sheet_name}' normalized to '{norm_name}'")

        if not actual_sheet_name:
            print(f"NO matching sheet found for FS meeting: {raw_sheet_name}")
            continue

        print(f"Matched FS sheet: {actual_sheet_name}")
        print(f"Horses in FS for this meeting: {list(horses.keys())}")

        sheet = workbook[actual_sheet_name]

        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                horse_name = str(cell.value).strip() if cell.value else ""
                if horse_name in horses:
                    fs_value = horses[horse_name]
                    sheet.cell(row=cell.row, column=23, value=fs_value)

                    print(f"   ➕ FS Saved | Row {cell.row} | Horse: '{horse_name}' | Value: {fs_value}")


    # --- PROCESS SKY RATING ---
    print("\n==============================")
    print("PROCESSING SKY RATING (Col X)")
    print("==============================")

    for raw_sheet_name, horses in SR.items():
        norm_name = normalize(raw_sheet_name)
        actual_sheet_name = normalized_sheet_map.get(norm_name)

        print(f"\n➡ Meeting SR: '{raw_sheet_name}' normalized to '{norm_name}'")

        if not actual_sheet_name:
            print(f"NO matching sheet found for SR meeting: {raw_sheet_name}")
            continue

        print(f"Matched SR sheet: {actual_sheet_name}")
        print(f"Horses in SR for this meeting: {list(horses.keys())}")

        sheet = workbook[actual_sheet_name]

        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                horse_name = str(cell.value).strip() if cell.value else ""
                if horse_name in horses:
                    sky_value = horses[horse_name]
                    sheet.cell(row=cell.row, column=24, value=sky_value)

                    print(f"   Sky Saved | Row {cell.row} | Horse: '{horse_name}' | Value: {sky_value}")

    workbook.save(excel_file)
    print("\n==============================")
    print("Excel updated successfully")
    print("==============================\n")



def main():
    
    driver = setup_driver()
    target_meetings = get_target_meetings_from_excel(FILE_NAME)
    get_meetings(driver=driver, url=BASE_URL + "/racing/meetings/today/", target_meetings=target_meetings)

    merge_excel(FILE_NAME, FS)

    driver.quit()


if __name__ == '__main__':
    main()
